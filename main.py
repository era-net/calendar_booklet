import os
from datetime import datetime, timedelta
import holidays
import openpyxl
import json
import csv
from fpdf import FPDF

''' Parameters - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - '''
year = 2022
date_format = "%m/%d/%y"
date_delimiter = str(date_format.split("%")[1][1])
holiday_country = "US" # Available Countries: https://github.com/dr-prodigy/python-holidays#available-countries
holiday_subdiv = "CA" # default = None
include_holidays = True
include_custom_events = True
include_page_no = False
language = "en"
''' - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - '''

def hex_to_rgb(hex):
    strp_hex = hex.lstrip("#")
    return tuple(int(strp_hex[i:i+2], 16) for i in (0, 2, 4))

# get translations
lang_exists = None
for lang_file in os.listdir("lang"):
    lang_split = lang_file.split(".")
    if lang_split[0] == language:
        language_file = "lang/"+lang_file
        lang_exists = True
        break
if not lang_exists:
    print("language is not existent ...")
    exit()

# load translations
with open(language_file, "r", encoding="utf-8") as lf:
    language_translation = json.loads(lf.read())

# generating main object
start = datetime(int(year), 1, 1)
end = datetime(int(year), 12, 31)

# generate general year_data dict
delta = end - start
year_data = {}
for n in range(12):
    year_data[str(n)] = {}
    for i in range(delta.days + 1):
        day = start + timedelta(days=i)
        month_num = int(day.month) - 1
        date = datetime.strftime(day, date_format)
        if month_num == n:
            year_data[str(n)][str(date)] = {"weekday": int(day.weekday()), "events": []}

# insert holydays
if include_holidays:
    default_colors = {
        "holidays": {
            "font": "#000000",
            "background": "#C5E0B3"
        }
    }
    if not os.path.isdir("inc"):
        os.mkdir("inc")
    if not os.path.isfile("inc/colors.json"):
        with open("inc/colors.json", "w", encoding="utf-8") as jf:
            jf.write(json.dumps(default_colors, indent=4))
    for x in holidays.country_holidays(holiday_country, years=year, subdiv=holiday_subdiv).items():
        mon_int = int(x[0].month) - 1
        year_data[str(mon_int)][x[0].strftime(date_format)]["events"].append({
            "name": x[1].upper(),
            "desc": None,
            "type": "holidays"
        })

# insert custom events
if include_custom_events:
    for file in os.listdir("events"):
        if file.split(".")[1] == "xlsx":
            xlsx = openpyxl.load_workbook("events/"+file)
            sheet_list = list(xlsx.active.values)
            if len(sheet_list) > 1:
                sheet_list.pop(0)
                for event in sheet_list:
                    event_name = event[0]
                    event_date_day = event[1]
                    event_date_month = int(event[2])
                    mon_int = event_date_month - 1
                    dtt = datetime(int(year), int(event_date_month), int(event_date_day))
                    date_str = dtt.strftime(date_format)
                    year_data[str(mon_int)][date_str]["events"].append({
                        "name": event_name,
                        "type": file
                    })
        elif file.split(".")[1] == "csv":
            csv_file = "events/"+file
            with open(csv_file, "r", encoding="utf-8") as cf:
                csv_reader = csv.reader(cf)
                next(csv_reader, None)
                for line in csv_reader:
                    event_name = str(line[0])
                    event_date_day = line[1]
                    event_date_month = int(line[2])
                    mon_int = event_date_month - 1
                    dtt = datetime(int(year), int(event_date_month), int(event_date_day))
                    date_str = dtt.strftime(date_format)
                    year_data[str(mon_int)][date_str]["events"].append({
                        "name": event_name,
                        "type": file
                    })
        else:
            print("please only use .xlsx and .csv files for your events")


month_strings = language_translation["month_strings"]
short_weekdays = language_translation["short_weekdays"]

pdf = FPDF()

# loading fonts
pdf.add_font("AnekLatin", "B", os.getcwd() + "\\fonts\\AnekLatin\\AnekLatin-Bold.ttf", uni=True)
pdf.add_font("Roboto", "", os.getcwd() + "\\fonts\\Roboto\\Roboto-Regular.ttf", uni=True)
pdf.add_font("Roboto", "B", os.getcwd() + "\\fonts\\Roboto\\Roboto-Bold.ttf", uni=True)

# get colors
with open("inc/colors.json", "r", encoding="utf-8") as cf:
    _colors = json.loads(cf.read())

# generate pdf
for mon in year_data:
    pdf.add_page()
    pdf.set_font('AnekLatin', 'B', 22)
    pdf.cell(0, 15, month_strings[int(mon)] + " " + str(year), 0, 1, 'L', 0)
    pdf.set_font('Roboto', 'B', 12)
    pdf.cell(pdf.w / 10.5, 8, language_translation["date"], 1, 0)
    pdf.cell(pdf.w / 10.5, 8, language_translation["hours"], 1, 0)
    pdf.cell(pdf.w / 10.5, 8, language_translation["currency"], 1, 0)
    pdf.cell(pdf.w / 1.6, 8, language_translation["description"], 1, 1)
    pdf.set_font('Roboto', '', 12)
    for date in year_data[mon]:
        # grabbing the date
        date_obj = datetime.strptime(date, date_format)
        print_day = date_obj.day
        print_month = date_obj.month
        if print_day < 10:
            print_day = "0" + str(print_day)
        if print_month < 10:
            print_month = "0" + str(print_month)
        print_date = str(print_day) + date_delimiter + str(print_month)
        pdf.cell(pdf.w / 10.5, 7.6, print_date, 1, 0, 'R')

        # grabbing weekday
        if year_data[mon][date]["weekday"] == 5 or year_data[mon][date]["weekday"] == 6:
            pdf.set_font('Roboto', 'B', 12)
            pdf.cell(pdf.w / 10.5, 7.6, short_weekdays[year_data[mon][date]["weekday"]], 1, 0, 'L')
        else:
            pdf.set_font('Roboto', '', 12)
            pdf.cell(pdf.w / 10.5, 7.6, "", 1, 0, 'L')
        
        # insert column to write the hours by hand
        pdf.set_font('Roboto', '', 12)
        pdf.cell(pdf.w / 10.5, 7.6, "", 1, 0, 'L')

        # description column
        if len(year_data[mon][date]["events"]) > 0:
            pdf.set_font('Roboto', 'B', 12)
            if len(year_data[mon][date]["events"]) > 1:
                evt_str = ""
                has_holidays = False
                for event in year_data[mon][date]["events"]:
                    if "holidays" in event["type"]:
                        has_holidays = True
                        break
                for i, event in enumerate(year_data[mon][date]["events"]):
                    if i+1 < len(year_data[mon][date]["events"]):
                        evt_str += event["name"] + " | "
                    else:
                        evt_str += event["name"]
                if has_holidays:
                    pdf.set_fill_color(197, 224, 179)
                else:
                    pdf.set_fill_color(255, 255, 255)
                pdf.set_text_color(255, 0, 0)
                pdf.cell(pdf.w / 1.6, 7.6, evt_str, 1, 1, "L", True)
                pdf.set_text_color(0, 0, 0)
            else:
                for event in year_data[mon][date]["events"]:
                    if event["type"] in _colors:
                        fill_color = hex_to_rgb(_colors[event["type"]]["background"])
                        font_color = hex_to_rgb(_colors[event["type"]]["font"])
                        pdf.set_fill_color(fill_color[0], fill_color[1], fill_color[2])
                        pdf.set_text_color(font_color[0], font_color[1], font_color[2])
                    else:
                        pdf.set_fill_color(255, 255, 255)
                        pdf.set_text_color(0, 0, 0)
                    pdf.cell(pdf.w / 1.6, 7.6, event["name"], 1, 1, "L", True)
                    pdf.set_fill_color(255, 225, 225)
                    pdf.set_text_color(0, 0, 0)
            pdf.set_font('Roboto', '', 12)
        else:
            pdf.set_font('Roboto', '', 12)
            pdf.cell(pdf.w / 1.6, 7.6, "", 1, 1)
    pdf.set_font('Roboto', 'U', 12)
    pdf.cell(pdf.w / 10.5, 7.6, "", 0, 0)
    pdf.cell(pdf.w / 10.5, 7.6, "                 ", 1, 0)
    pdf.cell(pdf.w / 10.5, 7.6, "                 ", 1, 0)
    pdf.set_font('Roboto', '', 12)
    if include_page_no:
        page = language_translation["site_prefix"] + " | " + str(pdf.page_no())
        str_wdt = pdf.get_string_width(page)
        pdf.text((pdf.w - 5) - str_wdt, pdf.h - 5, page)
pdf.output("booklet.pdf")